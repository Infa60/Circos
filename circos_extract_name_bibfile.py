"""
================================================================================
BIBTEX TO EXCEL SHORT REFERENCE EXTRACTOR
================================================================================

Description:
This script reads a `.bib` file (such as an export from Zotero, Mendeley, or
EndNote) and generates a clean Excel file mapping short citation references
to their full article titles.

Key Features:
1. Custom Parsing: Robustly parses BibTeX syntax without requiring heavy external
   dependencies like `bibtexparser`. It handles encoding fallbacks automatically.
2. Smart Author Formatting:
   - 1 author: "Author YYYY" (e.g., Smith 2020)
   - 2 authors: "A1 & A2 YYYY" (e.g., Doe & Lee 2021)
   - 3+ authors: "FirstAuthor et al. YYYY" (e.g., Brown et al. 2019)
3. Duplicate Handling: If multiple papers share the exact same author(s) and year,
   the script automatically appends suffixes ('a', 'b', 'c', etc.) to distinguish
   them (e.g., "Smith 2020a", "Smith 2020b").
4. Text Cleanup: Strips stray BibTeX formatting brackets (e.g., `{title}`) and
   normalizes excessive whitespace.

Output:
An Excel (`.xlsx`) file containing two columns:
  - ref: The short citation reference.
  - title: The cleaned article title.
================================================================================
"""

import re
from pathlib import Path
from typing import List, Dict, Tuple

# =================
# TO CUSTOMIZE:
# =================
BIB_PATH = r"C:\Users\bourgema\OneDrive - Université de Genève\Documents\ENABLE\Review\Review_code\Exported Items.bib"
OUT_XLSX = r"C:\Users\bourgema\OneDrive - Université de Genève\Documents\ENABLE\Review\Review_code\short_refs_with_titles.xlsx"

SORT_OUTPUT = True  # Final alphabetical sort (based on 'ref')
DROP_MISSING = True  # Ignore entries without author or year

# ================
# XLSX Dependency:
# ================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError as e:
    raise SystemExit(
        "The 'openpyxl' module is required to write .xlsx files.\n"
        "Install it in your PyCharm environment:\n"
        "    pip install openpyxl"
    )

# =================================
# Nothing to modify below this line
# =================================

ENTRY_START_RE = re.compile(r"@\s*(\w+)\s*\{\s*([^,]+)\s*,", re.UNICODE)


def read_text_smart(path: Path) -> str:
    """Robust reading of the .bib file with encoding fallback."""
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _strip_outer_quotes_or_braces(val: str) -> str:
    if not val:
        return ""
    v = val.strip()
    if (v.startswith("{") and v.endswith("}")) or (
        v.startswith('"') and v.endswith('"')
    ):
        return v[1:-1]
    return v


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _remove_outer_braces_from_name(name: str) -> str:
    # {de la Cruz} -> de la Cruz
    return re.sub(r"^\{(.+)\}$", r"\1", _norm_ws(name))


def clean_title(title: str) -> str:
    """Removes ALL braces, normalizes spaces."""
    t = _strip_outer_quotes_or_braces(title)
    t = re.sub(r"[{}]", "", t)  # removes all braces
    t = _norm_ws(t)
    return t


def parse_entries(text: str) -> List[str]:
    """Returns the list of text blocks for each BibTeX entry."""
    entries: List[str] = []
    i, n = 0, len(text)
    while True:
        m = ENTRY_START_RE.search(text, i)
        if not m:
            break
        start = m.start()
        j = text.find("{", start)
        depth = 0
        end = None
        for k in range(j, n):
            ch = text[k]
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = k + 1
                    break
        if end is None:
            break
        entries.append(text[start:end])
        i = end
    return entries


def parse_fields(entry_text: str) -> Dict[str, str]:
    """Extracts a field->value dict for a BibTeX entry (robustly)."""
    start = entry_text.find("{")
    if start < 0:
        return {}
    comma = entry_text.find(",", start)
    if comma < 0:
        return {}
    body = entry_text[comma + 1 : entry_text.rfind("}")]
    parts: List[str] = []
    buf: List[str] = []
    depth = 0
    inq = False
    for ch in body:
        if ch == "{" and not inq:
            depth += 1
            buf.append(ch)
        elif ch == "}" and not inq:
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch == '"' and depth == 0:
            inq = not inq
            buf.append(ch)
        elif ch == "," and depth == 0 and not inq:
            s = "".join(buf).strip()
            if s:
                parts.append(s)
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append("".join(buf).strip())

    out: Dict[str, str] = {}
    for p in parts:
        if "=" in p:
            k, v = p.split("=", 1)
            key = _norm_ws(k).lower()
            val = _strip_outer_quotes_or_braces(_norm_ws(v))
            out[key] = val
    return out


def split_authors(s: str) -> List[str]:
    if not s:
        return []
    # BibTeX separates authors with " and "
    return [_norm_ws(part) for part in s.split(" and ") if _norm_ws(part)]


def display_surname(author: str) -> str:
    """
    Returns the 'displayed' surname (accents preserved).
    Handles 'Last, First' or 'First Last' and removes enclosing braces.
    """
    a = _remove_outer_braces_from_name(author)
    if "," in a:
        sur = a.split(",", 1)[0]
    else:
        parts = a.split(" ")
        sur = parts[-1] if parts else ""
    return _norm_ws(sur)


def year_from_fields(f: Dict[str, str]) -> str:
    # Priority to 'year', otherwise 1st YYYY sequence in 'date'
    if f.get("year"):
        m = re.search(r"\d{4}", f["year"])
        if m:
            return m.group(0)
    if f.get("date"):
        m = re.search(r"\d{4}", f["date"])
        if m:
            return m.group(0)
    return ""


def short_ref_and_title(f: Dict[str, str]) -> Tuple[str, str]:
    authors = split_authors(f.get("author", ""))
    year = year_from_fields(f)
    title = clean_title(f.get("title", ""))
    if not authors or not year:
        return "", title
    if len(authors) >= 3:
        lead = display_surname(authors[0])
        ref = f"{lead} et al. {year}"
    elif len(authors) == 2:
        a1 = display_surname(authors[0])
        a2 = display_surname(authors[1])
        ref = f"{a1} & {a2} {year}"
    else:
        a1 = display_surname(authors[0])
        ref = f"{a1} {year}"
    return ref, title


def write_excel(rows: List[Tuple[str, str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "refs"

    # Headers
    ws.append(["ref", "title"])
    bold = Font(bold=True)
    ws["A1"].font = bold
    ws["B1"].font = bold

    # Rows
    for ref, title in rows:
        ws.append([ref, title])

    # Simple auto column width
    for col in ("A", "B"):
        max_len = max(
            (len(str(ws[f"{col}{r}"].value)) for r in range(1, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 80)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def main():
    bib_path = Path(BIB_PATH)
    out_path = Path(OUT_XLSX)

    if not bib_path.exists():
        raise FileNotFoundError(f"Bib file not found: {bib_path}")

    text = read_text_smart(bib_path)

    # Parse and count entries
    entry_blocks = parse_entries(text)
    print(f"📄 Entries read from the .bib: {len(entry_blocks)}")

    # Build base refs + titles
    base_rows: List[Tuple[str, str]] = []  # (ref_base, title)
    for block in entry_blocks:
        f = parse_fields(block)
        ref, title = short_ref_and_title(f)
        if ref:
            base_rows.append((ref, title))
        elif not DROP_MISSING:
            base_rows.append(("", title))

    # Total count per base ref
    totals: Dict[str, int] = {}
    for ref, _ in base_rows:
        if not ref:
            continue
        totals[ref] = totals.get(ref, 0) + 1

    # Assign a/b/c suffixes for all occurrences of a base ref
    seen_index: Dict[str, int] = {}
    final_rows: List[Tuple[str, str]] = []
    for ref, title in base_rows:
        if not ref:
            continue
        if totals.get(ref, 0) >= 2:
            idx = seen_index.get(ref, 0)  # 0,1,2...
            if idx < 26:
                suffix = chr(ord("a") + idx)  # a..z
            else:
                # beyond 26 → aa, ab, ...
                first = chr(ord("a") + (idx // 26) - 1)
                second = chr(ord("a") + (idx % 26))
                suffix = f"{first}{second}"
            final_rows.append((f"{ref}{suffix}", title))
            seen_index[ref] = idx + 1
        else:
            final_rows.append((ref, title))

    if SORT_OUTPUT:
        final_rows = sorted(final_rows, key=lambda x: x[0].lower())

    write_excel(final_rows, out_path)
    print(f"✅ {len(final_rows)} rows written to: {out_path}")


if __name__ == "__main__":
    main()
